from flask import Flask, render_template, request, redirect, url_for
import openpyxl
import os

app = Flask(__name__)

# Path file Excel
DATA_FOLDER = os.path.join(os.getcwd(), 'data')
os.makedirs(DATA_FOLDER, exist_ok=True)
excel_path = os.path.join(DATA_FOLDER, 'data_keuangan.xlsx')

# Buat file Excel kalau belum ada
if not os.path.exists(excel_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transaksi"
    ws.append(['Tanggal', 'Kategori', 'Tipe', 'Jumlah', 'Keterangan'])
    wb.save(excel_path)

# Hitung saldo dan total pemasukkan/pengeluaran
def hitung_keuangan():
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    saldo = 0
    total_pemasukkan = 0
    total_pengeluaran = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] == 'Pemasukkan':
            saldo += int(row[3])
            total_pemasukkan += int(row[3])
        elif row[2] == 'Pengeluaran':
            saldo -= int(row[3])
            total_pengeluaran += int(row[3])
    
    return saldo, total_pemasukkan, total_pengeluaran

@app.route('/')
def redirect_home():
    return redirect(url_for('form'))

@app.route('/form')
def form():
    return render_template('index.html')

@app.route('/hapus/<int:index>', methods=['GET'])
def hapus(index):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    ws.delete_rows(index + 2)
    wb.save(excel_path)
    return redirect(url_for('data'))

@app.route('/edit/<int:index>', methods=['GET', 'POST'])
def edit(index):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    if request.method == 'POST':
        tanggal = request.form['tanggal']
        kategori = request.form['kategori']
        tipe = request.form['tipe']
        jumlah = request.form['jumlah']
        keterangan = request.form['keterangan']
        
        # Update data
        row = index + 2
        ws.cell(row=row, column=1, value=tanggal)
        ws.cell(row=row, column=2, value=kategori)
        ws.cell(row=row, column=3, value=tipe)
        ws.cell(row=row, column=4, value=jumlah)
        ws.cell(row=row, column=5, value=keterangan)
        
        wb.save(excel_path)
        return redirect(url_for('data'))
    
    # GET method - tampilkan form edit
    row = index + 2
    data = {
        'tanggal': ws.cell(row=row, column=1).value,
        'kategori': ws.cell(row=row, column=2).value,
        'tipe': ws.cell(row=row, column=3).value,
        'jumlah': ws.cell(row=row, column=4).value,
        'keterangan': ws.cell(row=row, column=5).value
    }
    return render_template('edit.html', data=data, index=index)

@app.route('/tambah', methods=['POST'])
def tambah():
    tanggal = request.form['tanggal']
    kategori = request.form['kategori']
    tipe = request.form['tipe']
    jumlah = request.form['jumlah']
    keterangan = request.form['keterangan']

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    ws.append([tanggal, kategori, tipe, jumlah, keterangan])
    wb.save(excel_path)

    return redirect(url_for('data'))

@app.route('/data')
def data():
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    data = [{'index': i, 'tanggal': row[0], 'kategori': row[1], 'tipe': row[2], 'jumlah': row[3], 'keterangan': row[4]} 
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True))]
    
    saldo, total_pemasukkan, total_pengeluaran = hitung_keuangan()
    return render_template('data.html', 
                         data=data, 
                         saldo=saldo,
                         total_pemasukkan=total_pemasukkan,
                         total_pengeluaran=total_pengeluaran)

if __name__ == '__main__':
    app.run(debug=True)
